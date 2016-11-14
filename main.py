import sys
import os.path
from openpyxl import Workbook, load_workbook#, compat.range

def main():
    if(len(sys.argv) < 3):
        print('You must provide at least 2 spreadsheets to combine')
        print('\t{} <sheet1.xlsx> <sheet2.xlsx> [sheet3.xlsx ...]'.format(sys.argv[0]))
        sys.exit(1)

    output = Workbook()
    outputWs = output.create_sheet('combined',0)
    first = True
    rowTotal = 0
    writtenRows = []
    for fileName in sys.argv[1:]:
        if(not os.path.isfile(fileName)):
            print('Cannot find the file "{}". Check the name and try again'.format(fileName))
            sys.exit(2)
        wb = Workbook()
        wb = load_workbook(filename=fileName, read_only=True)
        skip=True
        skipped = 0
        for rowNum,r in enumerate(wb.active.rows, 1):
            if(skip and not first):
                skip=False
                skipped += 1
                continue
            skip=False
            cellsToWrite = []
            for cell in r:
                cellsToWrite.append(cell.value)
            if(cellsToWrite not in writtenRows):
                for i,value in enumerate(cellsToWrite, 1):
                    outputWs.cell(row=rowNum+rowTotal - skipped, column=i).value = value
                writtenRows.append(cellsToWrite)
            else:
                skipped += 1
        rowTotal += rowNum - skipped
        first = False
    output.save('output.xlsx')

if __name__ == '__main__':
    main()
